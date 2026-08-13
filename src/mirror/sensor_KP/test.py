import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl import Workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.comments import Comment
import re
from typing import Optional, List, Dict, Any
import time
from datetime import datetime
from pathlib import Path


class ExcelDataHandler:
    """Excel数据处理器，支持图片插入、排序和合并单元格"""
    
    def __init__(self, excel_path: str = "./mirror2_data/sensor_data.xlsx", 
                 all_actuator_info: Optional[Dict] = None, 
                 is_merge_cell: bool = False):
        """
        初始化Excel处理器
        
        Args:
            excel_path: Excel文件路径
            all_actuator_info: 执行器信息字典
            is_merge_cell: 是否合并单元格
        """
        self.excel_path = excel_path
        self.all_actuator_info = all_actuator_info or {}
        self.is_merge_cell = is_merge_cell
        
        # 表头列映射
        self.header_col_map = {
            "测试时间": "A", 
            "电机id": "B", 
            "传感器id": "C", 
            "弹簧id": "D", 
            "传感器量程": "E", 
            "测试区间": "F", 
            "脉冲范围": "G", 
            "拟合图": "H", 
            "线性方程": "I", 
            "线性度": "J", 
            "平均线性度": "K"
        }
        
        # 图片尺寸配置
        self.image_width = 120
        self.image_height = 120
        self.row_height = 90
        self.col_width = 18
        
        # 初始化Excel
        self.init_excel()
        self.start_row = self.ws.max_row + 1
        
        # 存储每行图片的锚点信息，用于排序
        self.row_image_anchors = {}

    def init_excel(self) -> None:
        """初始化Excel文件，如果不存在则创建"""
        # 确保目录存在
        Path(self.excel_path).parent.mkdir(parents=True, exist_ok=True)
        
        if not os.path.exists(self.excel_path):
            self._create_new_excel()
        else:
            try:
                self.wb = load_workbook(self.excel_path, data_only=False)
                if "Sensor Data" not in self.wb.sheetnames:
                    self.ws = self.wb.create_sheet("Sensor Data")
                else:
                    self.ws = self.wb["Sensor Data"]
                
                # 清除旧的图片引用，避免内存泄漏
                self.ws._images = []
                
            except Exception as e:
                print(f"⚠️ 检测到Excel文件损坏，自动备份并重建: {str(e)}")
                self._backup_and_recreate()

    def _create_new_excel(self) -> None:
        """创建新的Excel文件"""
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Sensor Data"
        
        # 添加表头
        headers = ["测试时间", "电机id", "传感器id", "弹簧id", "传感器量程", 
                   "测试区间", "脉冲范围", "拟合图", "线性方程", "线性度", "平均线性度"]
        self.ws.append(headers)
        
        # 设置表头样式
        for col in range(1, len(headers) + 1):
            cell = self.ws.cell(row=1, column=col)
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        self.wb.save(self.excel_path)

    def _backup_and_recreate(self) -> None:
        """备份损坏文件并重新创建"""
        backup_path = f"{self.excel_path}.broken_bak_{int(time.time())}.xlsx"
        os.rename(self.excel_path, backup_path)
        print(f"📁 损坏文件已备份到: {backup_path}")
        self._create_new_excel()

    def _create_image_with_anchor(self, image_path: str, cell_ref: str) -> Optional[Image]:
        """
        创建带双单元格锚点的图片
        
        Args:
            image_path: 图片路径
            cell_ref: 单元格位置，如 'H3'
            
        Returns:
            带锚点的图片对象，失败返回None
        """
        if not os.path.exists(image_path):
            return None
        
        try:
            # 解析单元格位置
            col_letter = ''.join(filter(str.isalpha, cell_ref))
            row_num = int(''.join(filter(str.isdigit, cell_ref)))
            col_idx = column_index_from_string(col_letter)
            
            # 加载图片
            img = Image(image_path)
            
            # 计算缩放比例
            orig_w, orig_h = img.width, img.height
            scale = min(self.image_width / orig_w, self.image_height / orig_h)
            img.width = int(orig_w * scale)
            img.height = int(orig_h * scale)
            
            # 创建双单元格锚点（TwoCellAnchor）
            # 左上角锚点：目标单元格的左上角
            from_anchor = AnchorMarker(col=col_idx - 1, colOff=0, row=row_num - 1, rowOff=0)
            
            # 右下角锚点：目标单元格的右下角（根据图片大小动态计算）
            # 使用目标单元格的右下角作为锚点，确保图片在单元格内
            to_anchor = AnchorMarker(
                col=col_idx - 1, 
                colOff=int(img.width * 9525),  # 转换为EMU单位（1像素 ≈ 9525 EMU）
                row=row_num - 1, 
                rowOff=int(img.height * 9525)
            )
            
            # 创建双单元格锚点
            img.anchor = TwoCellAnchor('twoCell', from_anchor, to_anchor)
            
            # 存储锚点信息用于排序
            self.row_image_anchors[row_num] = {
                'image': img,
                'cell_ref': cell_ref,
                'row_num': row_num
            }
            
            return img
            
        except Exception as e:
            print(f"❌ 创建图片锚点失败: {e}")
            return None

    def insert_image_to_cell(self, cell_ref: str, image_path: str, 
                             img_name: str = "") -> bool:
        """
        插入图片到指定单元格（使用双单元格锚点）
        
        Args:
            cell_ref: 单元格位置，如 'H3'
            image_path: 图片路径
            img_name: 图片名称（将写入单元格）
            
        Returns:
            是否插入成功
        """
        if not os.path.exists(image_path):
            print(f"⚠️ 图片不存在: {image_path}")
            return False
        
        try:
            # 创建带锚点的图片
            img = self._create_image_with_anchor(image_path, cell_ref)
            if img is None:
                return False
            
            # 插入图片
            self.ws.add_image(img)
            
            # 设置单元格内容（图片名称）
            target_cell = self.ws[cell_ref]
            target_cell.value = img_name or os.path.basename(image_path)
            target_cell.alignment = Alignment(
                horizontal='center', 
                vertical='bottom'
            )
            target_cell.font = Font(size=9, color="333333")
            
            # 自动调整行高列宽
            row_num = int(''.join(filter(str.isdigit, cell_ref)))
            col_letter = ''.join(filter(str.isalpha, cell_ref))
            self.ws.row_dimensions[row_num].height = self.row_height
            self.ws.column_dimensions[col_letter].width = self.col_width
            
            print(f"✅ 图片已插入到 {cell_ref}: {os.path.basename(image_path)}")
            return True
            
        except Exception as e:
            print(f"❌ 插入图片失败: {e}")
            return False

    def batch_insert_images(self) -> None:
        """批量插入图片"""
        current_row = self.start_row
        print(f"📊 当前表格总行数: {self.ws.max_row}，起始行: {current_row}")
        
        for img_name, actuator_info in self.all_actuator_info.items():
            full_img_path = actuator_info.get("拟合图", "")
            
            if not os.path.exists(full_img_path):
                print(f"⚠️ 跳过不存在的文件: {full_img_path}")
                continue
            
            # 构建单元格位置
            cell_ref = f"{self.header_col_map['拟合图']}{current_row}"
            
            # 插入图片（使用双单元格锚点）
            success = self.insert_image_to_cell(
                cell_ref=cell_ref,
                image_path=full_img_path,
                img_name=img_name
            )
            
            if success:
                # 写入其他数据
                self._write_row_data(current_row, actuator_info)
                current_row += 1
        
        # 保存Excel
        self.wb.save(self.excel_path)
        print(f"✅ 批量插入完成，共处理 {len(self.all_actuator_info)} 张图片")

    def _write_row_data(self, row: int, actuator_info: Dict) -> None:
        """写入行数据"""
        field_map = {
            "电机id": "电机id",
            "测试时间": "测试时间",
            "传感器id": "传感器id",
            "弹簧id": "弹簧id",
            "传感器量程": "传感器量程",
            "测试区间": "测试区间",
            "脉冲范围": "脉冲范围",
            "线性方程": "线性方程",
            "线性度": "线性度"
        }
        
        for col_key, field_key in field_map.items():
            if field_key in actuator_info:
                col_letter = self.header_col_map[col_key]
                self.ws[f"{col_letter}{row}"] = actuator_info[field_key]

    def sort_excel_with_images(self, start_row: int = 2) -> None:
        """
        对Excel数据进行排序（图片随行移动）
        
        Args:
            start_row: 数据起始行（1为表头）
        """
        if start_row > self.ws.max_row:
            print("⚠️ 没有数据需要排序")
            return
        
        print("🔄 开始排序数据...")
        
        # 1. 提取所有行的数据和图片
        rows_data = []
        for row in range(start_row, self.ws.max_row + 1):
            # 获取排序键
            motor_id = self.ws.cell(
                row=row, 
                column=column_index_from_string(self.header_col_map['电机id'])
            ).value
            test_time = self.ws.cell(
                row=row, 
                column=column_index_from_string(self.header_col_map['测试时间'])
            ).value
            
            # 获取整行所有单元格的值
            row_cells = []
            for col in range(1, self.ws.max_column + 1):
                row_cells.append(self.ws.cell(row=row, column=col).value)
            
            # 获取该行的图片锚点信息
            image_info = self.row_image_anchors.get(row)
            
            rows_data.append({
                'motor_id': motor_id,
                'test_time': test_time,
                'row_cells': row_cells,
                'original_row': row,
                'image_info': image_info
            })
        
        # 2. 排序
        rows_data.sort(key=lambda x: (x['motor_id'] or 0, x['test_time'] or ""), 
                      reverse=False)
        
        # 3. 构建新的图片锚点映射
        new_image_anchors = {}
        
        # 4. 写入排序后的数据
        for new_row_idx, row_data in enumerate(rows_data, start=start_row):
            # 写入数据
            for col_idx, cell_value in enumerate(row_data['row_cells'], start=1):
                self.ws.cell(row=new_row_idx, column=col_idx).value = cell_value
            
            # 如果有图片，重新创建锚点
            if row_data['image_info']:
                img_info = row_data['image_info']
                # 更新图片的锚点到新行
                old_cell_ref = img_info['cell_ref']
                old_col_letter = ''.join(filter(str.isalpha, old_cell_ref))
                new_cell_ref = f"{old_col_letter}{new_row_idx}"
                
                # 重新创建图片锚点
                img = img_info['image']
                new_img = self._create_image_with_anchor(
                    img.path,  # 注意：这里需要获取原始图片路径
                    new_cell_ref
                )
                if new_img:
                    self.ws.add_image(new_img)
                    new_image_anchors[new_row_idx] = {
                        'image': new_img,
                        'cell_ref': new_cell_ref,
                        'row_num': new_row_idx
                    }
        
        # 更新图片锚点映射
        self.row_image_anchors = new_image_anchors
        
        # 5. 合并单元格（如果需要）
        if self.is_merge_cell:
            self._merge_cells(start_row)
        
        # 6. 保存
        self.wb.save(self.excel_path)
        print(f"✅ 排序完成，共处理 {len(rows_data)} 行数据，图片已跟随移动")

    def _merge_cells(self, start_row: int) -> None:
        """合并指定列的单元格（每3行合并一次）"""
        merge_cols = ["电机id", "传感器id", "弹簧id", "传感器量程", "平均线性度"]
        
        for row in range(start_row, self.ws.max_row + 1, 3):
            # 确保不超出最大行数
            end_row = min(row + 2, self.ws.max_row)
            
            for col_key in merge_cols:
                col_letter = self.header_col_map[col_key]
                start_cell = f"{col_letter}{row}"
                end_cell = f"{col_letter}{end_row}"
                
                try:
                    self.ws.merge_cells(f"{start_cell}:{end_cell}")
                    v = self.ws[f"{self.header_col_map['拟合图']}{row}"].value
                    print(f"v_拟合图 = {v}")
                except Exception as e:
                    print(f"⚠️ 合并单元格 {start_cell}:{end_cell} 失败: {e}")
            
            # 设置平均线性度公式
            linearity_col = self.header_col_map["线性度"]
            avg_col = self.header_col_map["平均线性度"]
            self.ws[f"{avg_col}{row}"] = (
                f"=AVERAGE({linearity_col}{row}:{linearity_col}{end_row})"
            )
            
            # 居中显示
            for col_key in merge_cols:
                col_letter = self.header_col_map[col_key]
                self.ws[f"{col_letter}{row}"].alignment = Alignment(
                    horizontal='center', 
                    vertical='center'
                )

    def add_comment_to_cell(self, cell_ref: str, comment_text: str, 
                           author: str = "System") -> None:
        """添加批注"""
        if not comment_text:
            return
        
        comment = Comment(comment_text, author)
        self.ws[cell_ref].comment = comment

    def get_row_count(self) -> int:
        """获取数据行数"""
        return self.ws.max_row - 1  # 减去表头

    def clear_data(self, keep_header: bool = True) -> None:
        """清空数据"""
        start_row = 2 if keep_header else 1
        self.ws.delete_rows(start_row, self.ws.max_row - start_row + 1)
        self.row_image_anchors = {}
        self.wb.save(self.excel_path)
        print("✅ 数据已清空")

    def main(self) -> None:
        """主执行方法"""
        print("="*60)
        print("📊 Excel数据处理器开始运行")
        print(f"📁 文件路径: {self.excel_path}")
        print(f"📈 数据行数: {len(self.all_actuator_info)}")
        print("="*60)
        
        # 1. 批量插入图片
        self.batch_insert_images()
        
        # 2. 排序
        self.sort_excel_with_images()
        
        print("="*60)
        print("✅ 所有操作完成！")
        print("="*60)


# ==================== 使用示例 ====================

if __name__ == "__main__":
    # 准备测试数据
    all_actuator_info = {
        "motor1_data_20260812_154103_拟合图": {
            "传感器id": "192.168.0.100:1",
            "pmac_ip": "192.168.0.200",
            "测试区间": "[-150N, 150N]",
            "测试时间": "20260812_154103",
            "传感器量程": "200N",
            "脉冲范围": "[0, -20W, -5K]",
            "电机id": 1,
            "弹簧id": 1,
            "线性度": 4008,
            "amplifier_id": 0,
            "channel_id": 3,
            "mirror_id": 2,
            "sensor_index": 2,
            "拟合图": "./mirror1_data/motor1_data_20260812_154103_拟合图.png",
            "线性方程": "y=4008x+1"
        },
        "motor1_data_20260812_154626_拟合图": {
            "传感器id": "192.168.0.100:1",
            "pmac_ip": "192.168.0.200",
            "测试区间": "[-150N, 150N]",
            "测试时间": "20260812_154626",
            "传感器量程": "200N",
            "脉冲范围": "[0, -20W, -5K]",
            "电机id": 1,
            "弹簧id": 1,
            "线性度": 4009,
            "amplifier_id": 0,
            "channel_id": 3,
            "mirror_id": 2,
            "sensor_index": 2,
            "拟合图": "./mirror1_data/motor1_data_20260812_154626_拟合图.png",
            "线性方程": "y=4009x+1"
        },
        "motor1_data_20260812_155925_拟合图": {
            "传感器id": "192.168.0.100:1",
            "pmac_ip": "192.168.0.200",
            "测试区间": "[-150N, 150N]",
            "测试时间": "20260812_155925",
            "传感器量程": "200N",
            "脉冲范围": "[0, -20W, -5K]",
            "电机id": 1,
            "弹簧id": 1,
            "线性度": 4018,
            "amplifier_id": 0,
            "channel_id": 3,
            "mirror_id": 2,
            "sensor_index": 2,
            "拟合图": "./mirror1_data/motor1_data_20260812_155925_拟合图.png",
            "线性方程": "y=4018x+1"
        }
    }
    
    # 创建处理器实例并运行
    handler = ExcelDataHandler(
        excel_path="./mirror3_data/sensor_data.xlsx",
        all_actuator_info=all_actuator_info,
        is_merge_cell=True
    )
    handler.main()

    file = '/home/mustocs/hs/src/mirror/sensor_KP/mirror1_data/motor8_data_20260812_155925_拟合图.png'
    dir = os.path.dirname(file)
    print(f"dir = {dir}")