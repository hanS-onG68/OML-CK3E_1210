import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font
from openpyxl import Workbook
from openpyxl.utils import column_index_from_string
import re
from typing import Optional, List
import time

class ExcelDataHandler:
    def __init__(self, excel_path="./mirror2_data/sensor_data.xlsx", all_actuator_info: Optional[dict] =None, is_merge_cell=False):
        self.excel_path = excel_path
        self.init_excel()
        self.start_row = self.ws.max_row + 1
        self.all_actuator_info = all_actuator_info
        self.is_merge_cell = is_merge_cell
        self.header_col_map = { "测试时间": "A", "电机id": "B", "传感器id": "C", "弹簧id": "D", "传感器量程": "E", "测试区间": "F", "脉冲范围": "G", "拟合图": "H", "线性方程": "I", "线性度": "J", "平均线性度": "K" }
        pass

    def init_excel(self):
        if not os.path.exists(self.excel_path):
            os.makedirs(os.path.dirname(self.excel_path), exist_ok=True)
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "Sensor Data"
            # 添加表头
            self.ws.append(["测试时间", "电机id", "传感器id", "弹簧id", "传感器量程", "测试区间", "脉冲范围", "拟合图", "线性方程", "线性度", "平均线性度"])
            self.wb.save(self.excel_path)  # 保存文件到当前路径下
        else:
            try:
                self.wb = load_workbook(self.excel_path, data_only=False)  # 加载现有文件，data_only=True确保读取公式的计算结果而不是公式本身
                self.ws = self.wb["Sensor Data"]  # 选择工作表
                self.ws._images = []              # 彻底清空所有缓存的旧图片对象，丢弃全部失效的文件句柄
            except Exception as e:
                print(f"检测到Excel文件损坏, 自动备份后新建文件，错误信息：{str(e)}")
                # 破损文件自动重命名备份，不会丢之前的残留内容
                backup_path = self.excel_path + f".broken_bak_{int(time.time())}.xlsx"
                os.rename(self.excel_path, backup_path)
                # 自动生成全新合法工作簿
                self.wb = Workbook()
                self.ws = self.wb.active
                self.ws.title = "Sensor Data"
                self.ws.append(["测试时间", "电机id", "传感器id", "弹簧id", "传感器量程", "测试区间", "脉冲范围", "拟合图", "线性方程", "线性度", "平均线性度"])
                self.wb.save(self.excel_path)

    def batch_insert_images(self):
        current_row = self.start_row
        print(f"当前表格总行数{self.ws.max_row}，下一条新数据写入起始行：{current_row}")
        for img_name, actuator_info in self.all_actuator_info.items():
            full_img_path = actuator_info["拟合图"]
            if not os.path.exists(full_img_path):
                print(f"跳过不存在的文件：{full_img_path}")
                continue

            # 1.加载图片，统一设置尺寸避免变形
            img = Image(full_img_path)
            cell_pos = f"{self.header_col_map['拟合图']}{current_row}"  # 目标单元格位置
            self.ws.add_image(img, cell_pos)  # 图片锚定到单元格中：
            target_cell = self.ws[cell_pos]   # 目标单元格

            
            # 2.统一缩放到120x120像素，可根据你的单元格大小调整
            original_w, original_h = img.width, img.height
            scale = min(120/original_w, 120/original_h)
            img.width = int(original_w * scale)
            img.height = int(original_h * scale)
            # 自动调整对应行高和列宽，适配图片大小
            self.ws.row_dimensions[current_row].height = 90
            self.ws.column_dimensions[self.header_col_map['拟合图']].width = 18

            # 3. 直接在同一个单元格写入文件名，自动靠下居中对齐
            target_cell.value = img_name
            target_cell.alignment = Alignment(horizontal='center', vertical='bottom')  # 设置文字样式：底部居中，字号调小一点，不遮挡图片
            target_cell.font = Font(size=9, color="333333")

            # 4. 将其他信息写入对应的单元格
            self.ws[f"{self.header_col_map['电机id']}{current_row}"] = actuator_info["电机id"]
            self.ws[f"{self.header_col_map['测试时间']}{current_row}"] = actuator_info["测试时间"]
            self.ws[f"{self.header_col_map['传感器id']}{current_row}"] = actuator_info["传感器id"]
            self.ws[f"{self.header_col_map['弹簧id']}{current_row}"] = actuator_info["弹簧id"]
            self.ws[f"{self.header_col_map['传感器量程']}{current_row}"] = actuator_info["传感器量程"]
            self.ws[f"{self.header_col_map['测试区间']}{current_row}"] = actuator_info["测试区间"]
            self.ws[f"{self.header_col_map['脉冲范围']}{current_row}"] = actuator_info["脉冲范围"]
            self.ws[f"{self.header_col_map['线性方程']}{current_row}"] = actuator_info["线性方程"]
            self.ws[f"{self.header_col_map['线性度']}{current_row}"] = actuator_info["线性度"]

            current_row += 1
        self.wb.save(self.excel_path)
        print(f"批量插入完成，共处理{len(self.all_actuator_info)}张图片")


    def sort_excel_with_images(self, start_row=2):
        # 1. 先把所有行的数据+图片锚点信息读取出来
        rows_data = []
        for row in range(start_row, self.ws.max_row + 1):
            a_col_val = self.ws.cell(row=row, column=column_index_from_string(self.header_col_map['电机id'])).value      # 排序优先级1：(电机id)列作为第一排序键
            b_col_val = self.ws.cell(row=row, column=column_index_from_string(self.header_col_map['测试时间'])).value    # 排序优先级2：(测试时间)列作为第二排序键
            row_cells = [self.ws.cell(row=row, column=c).value for c in range(1, self.ws.max_column+1)]                 # 把整行所有单元格的值、样式、图片锚点全部打包
            rows_data.append( (a_col_val, b_col_val, row_cells) )

        # 2. 按排序键排序: 自动兼容数字、日期、字符串类型
        rows_data.sort(key=lambda x: (x[0], x[1]), reverse=False)  # 按(电机id)列升序排序，如果(电机id)列相同则按(测试时间)列升序排序

        # 3. 把排序后的数据写回表格
        for new_row_idx, (a_key, b_key, row_cells) in enumerate(rows_data, start=start_row):
            for col_idx, cell_value in enumerate(row_cells, start=1):
                self.ws.cell(row=new_row_idx, column=col_idx).value = cell_value

        # 4. 合并单元格
        if self.is_merge_cell:
            for row in range(start_row, self.ws.max_row + 1, 3):
                self.ws.merge_cells(f"{self.header_col_map['电机id']}{row}:{self.header_col_map['电机id']}{row + 2}")
                self.ws.merge_cells(f"{self.header_col_map['传感器id']}{row}:{self.header_col_map['传感器id']}{row + 2}")
                self.ws.merge_cells(f"{self.header_col_map['弹簧id']}{row}:{self.header_col_map['弹簧id']}{row + 2}")
                self.ws.merge_cells(f"{self.header_col_map['传感器量程']}{row}:{self.header_col_map['传感器量程']}{row + 2}")
                self.ws.merge_cells(f"{self.header_col_map['平均线性度']}{row}:{self.header_col_map['平均线性度']}{row + 2}")
                self.ws[f"{self.header_col_map['平均线性度']}{row}"] = f"=AVERAGE({self.header_col_map['线性度']}{row}:{self.header_col_map['线性度']}{row + 2})"

        self.wb.save(self.excel_path)
        print("带图片的表格排序完成，所有图片自动跟随行移动")

    def main(self):
        self.batch_insert_images()
        self.sort_excel_with_images()


# 调用示例，适配你当前的sensor_KP项目
if __name__ == "__main__":
    all_actuator_info ={
        "motor1_data_20260729_144540_拟合图": {
            "传感器id": "192.168.0.100:1", 
            "pmac_ip": "192.168.0.200", 
            "测试区间": "[-150N, 150N]",
            "测试时间": "20260729_144540",
            "传感器量程": "200N",
            "脉冲范围": "[0, -20W, -5K]",
            "电机id": 1,
            "弹簧id": 1,
            "线性度": 4008,
            "amplifier_id": 0,
            "channel_id": 3,
            "mirror_id": 2,
            "sensor_index":2,   # 表示传感器在全部150个传感器中的索引位置，0~149
            "拟合图": "./mirror2_data/motor1_data_20260729_144540_拟合图.png",
            "线性方程": "y=4008x+1"
        },
        "motor1_data_20260729_153223_拟合图": {
                    "传感器id": "192.168.0.100:1", 
                    "pmac_ip": "192.168.0.200", 
                    "测试区间": "[-150N, 150N]",
                    "测试时间": "20260729_153223",
                    "传感器量程": "200N",
                    "脉冲范围": "[0, -20W, -5K]",
                    "电机id": 1,
                    "弹簧id": 1,
                    "线性度": 4009,
                    "amplifier_id": 0,
                    "channel_id": 3,
                    "mirror_id": 2,
                    "sensor_index":2,   # 表示传感器在全部150个传感器中的索引位置，0~149
                    "拟合图": "./mirror2_data/motor1_data_20260729_153223_拟合图.png",
                    "线性方程": "y=4009x+1"
        },
        "motor1_data_20260729_154924_拟合图": {
                    "传感器id": "192.168.0.100:1", 
                    "pmac_ip": "192.168.0.200", 
                    "测试区间": "[-150N, 150N]",
                    "测试时间": "20260729_154924",
                    "传感器量程": "200N",
                    "脉冲范围": "[0, -20W, -5K]",
                    "电机id": 1,
                    "弹簧id": 1,
                    "线性度": 4018,
                    "amplifier_id": 0,
                    "channel_id": 3,
                    "mirror_id": 2,
                    "sensor_index":2,   # 表示传感器在全部150个传感器中的索引位置，0~149
                    "拟合图": "./mirror2_data/motor1_data_20260729_154924_拟合图.png",
                    "线性方程": "y=4018x+1"
        }
    }
    ExcelDataHandler(excel_path="./mirror2_data/sensor_data.xlsx", all_actuator_info=all_actuator_info, is_merge_cell=True).main()
    pass
